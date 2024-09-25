from pathlib import Path
import re

import pandas as pd

from openpyxl import load_workbook

file_path = 'files/SEI_DNIT - 16453831 - Despacho (DNIT) - Copia.xlsx'
# file_path = 'files/SEI_DNIT - 16453831 - Despacho (DNIT).xlsx'

def extrair_sei(file_path_target = None, all_data = True):
    if file_path_target is not None:
        file_path = 'files/' + file_path_target
    try:
        df = pd.read_excel(file_path, sheet_name='Table 1')
    except ValueError:
        try:
            df = pd.read_excel(file_path, sheet_name='Tabela 1')
        except Exception as e:
            try:
                df = pd.read_excel(file_path, sheet_name='Planilha 1')
            except Exception as e:
                try:
                    df = pd.read_excel(file_path, sheet_name='Planilha1')
                except Exception as e:
                    print(f"Erro ao ler o arquivo: {e}")
                    return pd.DataFrame()
            

    if all_data:
        return df
    
    if 'TERMO DE ACEITE' in df.columns:
        def extract_sei_number(text):
            if text is None or pd.isna(text):
                return None
            try:
                match = re.search(r'\b\d{7}\b', str(text))
                if match:
                    return match.group()
                else:
                    match = re.search(r'\b\d{8}\b', str(text))
                    if match:
                        return match.group()
                return None
            except Exception as e:
                print(f"Erro na extração do SEI: {e}")
                return None

        df['SEI Number'] = df['TERMO DE ACEITE'].apply(extract_sei_number)

        df['Row Number'] = df.index + 2  # Soma 2 para ajustar ao número da linha no Excel (conta 1 para header e 1 para índice 0)

        filtered_df = df.dropna(subset=['SEI Number'])

        return filtered_df
    else:
        return pd.DataFrame()  # Retorna um DataFrame vazio se a coluna não existir

def realizar_validacao_excel(df, dictionary, file):
    
    #workbook = load_workbook(file_path)
    file_path = Path("./files/",file)
    workbook = load_workbook(file_path)
    
    if 'Table 1' in workbook.sheetnames:
        sheet = workbook['Table 1']
    elif 'Tabela 1' in workbook.sheetnames:
        sheet = workbook['Tabela 1']
    elif 'Tabela 1' in workbook.sheetnames:
        sheet = workbook['Planilha 1']
    else:
        print("Nenhuma planilha 'Table 1' ou 'Tabela 1' encontrada.")
        return

    # Determina o índice da coluna onde a nova coluna será inserida
    last_col = sheet.max_column + 1
    sheet.cell(row=1, column=last_col).value = "Validação"

    # Percorre as linhas e insere o valor "OK" ou "Não validado" na nova coluna
    for index, row in df.iterrows():
        validation_value = '-'
        if index in dictionary:
            validation_value = 'OK' if dictionary[index] else 'Inválido'
        
        sheet.cell(row=row['Row Number'], column=last_col).value = validation_value

    # Salva o workbook com as alterações
    workbook.save(file_path)
    print(f"Coluna 'Validação' adicionada e arquivo salvo em: {file_path}")